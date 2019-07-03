from trello import TrelloClient
from openpyxl import load_workbook
from copy import copy

import datetime
import re

blessID = '5bff57cb3b4c0b21262eabc6'
owenID = '5bff594d317049461251056e'
claraID = '54b3636bb2b9ce3343113869'
wilsonID = '5c2c5edca9970548e6e9b78e'
sunnyID = '5adc58e6bed38687f94a7fb5'
howardID = '5c6ce68c093e2182061dceb3'
maxID = '5c9b0ab20a49cc6f877b10ae'
whiteID = '5a6d4a5185c4f78a053fb5b1'
 
members = [blessID, owenID, claraID, wilsonID,
           whiteID, sunnyID, howardID, maxID]

key = 'd19356dd8fdeb53226fc5e3fd4f3ab1b'
secret = 'b8f25dfe09f13a2ab3318ed14692a1d8f8aad9a1db4c5183021ae9aa5b77fa23'

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


def calPercentage(card):
    if not card.checklists:
        return '0%'

    checklist = card.checklists[0].items    
    checked = 0
    for item in checklist:
        if item['checked']:
            checked = checked + 1

    if checked:
        return '{}%'.format((checked / len(checklist)) * 100)
    else:
        return '0%'


def findInAllBoards(client, nameList, id, percentage):
    print(id, nameList)
    li = []
    boards = client.list_boards()
    for board in boards:
        for lis in board.list_lists():
            if lis.name in nameList:
                for card in lis.list_cards():
                    print(card, card.member_id)
                    if id in card.member_id:

                        if percentage:
                            if lis.name == '進行中':
                                percent = calPercentage(card)
                            else:
                                percent = '100.00%'
                        else:
                            percent = ''

                        item = [
                            board.name,
                            '#{}'.format(card.short_id),
                            card.name,
                            percent
                        ]
                        print(item)
                        li.append(item)
    return li


def getAllItems(lists, percentage):
    client = TrelloClient(api_key=key, api_secret=secret)
    dones = []
    for memberId in members:
        done = findInAllBoards(client, lists, memberId, percentage)
        dones.append(done)
    return dones


def putInXlsx(wb, items, row):
    today = datetime.date.today()
    rangeDate = datetime.timedelta(days=6)
    startDate = today - rangeDate
    formattedDate = '日 期: {} ~ {}'.format(
        startDate.strftime('%Y/%m/%d'),
        today.strftime('%m/%d'),
    )
    for i in range(len(members)):
        ws = wb.worksheets[i]
        done = items[i]
        cell = copy(ws.cell(row=row, column=1))

        ws.cell(row=2, column=5).value = formattedDate

        for item in done:
            print(item)
            ws.insert_rows(row)
            for i in range(1, 6):
                ws.cell(row=row, column=i).font = copy(cell.font)
                ws.cell(row=row, column=i).border = copy(cell.border)
                ws.cell(row=row, column=i).fill = copy(cell.fill)
                ws.cell(row=row, column=i).number_format = copy(cell.number_format)
                ws.cell(row=row, column=i).protection = copy(cell.protection)
                # ws.cell(row=10, column=i).alignment = copy(cell.alignment)

            for i in range(1, 5):
                content = ILLEGAL_CHARACTERS_RE.sub(r'', item[i - 1])
                ws.cell(row=row, column=i).value = content
    return wb


def run():
    dones = getAllItems(['進行中', '已修改', '已上傳待複驗', '複驗完成'], True)
    inProgress = getAllItems(['進行中'], False)

    filepath = "./report.xlsx"
    wb = load_workbook(filepath)

    wb = putInXlsx(wb, inProgress, 10)
    wb = putInXlsx(wb, dones, 5)

    today = datetime.date.today()
    filename = "./工作週報_Frontend_{}.xlsx".format(
        today.strftime('%Y%m%d')
    )
    wb.save(filename=filename)

    print('done')


if __name__ == '__main__':
    run()
